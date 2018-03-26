#Equation Parser for python and Excel code
#Function Libaray
#by Regan Lu

#import libaray functions
from math import *
from sympy import simplify, sstrrepr, latex, preview, Symbol 
import pageConfig as pC
from openpyxl.utils.cell import _get_column_letter as gCol
from openpyxl as load_workbook
#import re


#local variables

#defining the excel function list
pyNameList = ['fabs(','factorial(','ceil(','log(','pow(','**' ]
xlsNameList = ['abs(','fact(','ceiling(','ln(','power(','^']

#Functions
    
#convert equation from processed sympy equation
#to excel friendly equation
def xlsfuncFind(String):
    #temp string
    tString = String
    
    #index for names
    i = 0
    
    #Excel Function Replacement Loop
    #Limited libaray replacement, further development will be required
    #Should be good enough for most application anyways
    for x in pyNameList:
        tString = tString.replace(x,xlsNameList[i])
        i = i + 1
    
    return tString

def xlsCell(varList, seqList, String):
    #temp string
    tString = String

    #Variable Replacement loops
    #variable list
    for x in varList:
        tString = tString.replace(x.var,x.cell)

    #sequence list
    for x in seqList:
        tString = tString.replace(x.var,x.cell)
    
    return tString

def Assigner(varList, seqList, iList):
    cRow = pC.tblock + pC.SizeY + 1
    page = 1

    cRow,page = vCellAssign(varList,cRow,page)

def sCellAssign(seqList,cRow,page):
    crow = cRow
    pg = page

    for x in seqList:
        crow = crow + pC.spacing

        if(x.type = '0'):
            if((crow+fSize)<(pg*pC.SizeY):
               x.cell = 'D' + str(crow-pC.rEqt)
               crow = crow+fSize
            else:
                pg = pg + 1
                crow = (pg)*pC.SizeY + pC.tblock
                x.cell = 'D' + str(crow-pC.rEqt)
                crow = crow + pC.fSize
        elif(x.type = '1'):
            if((crow+ifSize)<(pg*pC.SizeY):
               x.cell = 'D' + str(crow-pC.cift)
               crow = crow+fSize
            else:
                pg = pg + 1
                crow = (pg)*pC.SizeY + pC.tblock
                x.cell = 'D' + str(crow-pC.cift)
                crow = crow + pC.ifSize
        else:
               print("Error: Incorret Type")
def iCellAssign(iList,cRow,page):
    crow = cRow
    pg = page
    
    for x in iList:
        path = x.path
        f = load_workbook(path)
        mrow = max_row(f.active)
        crow = crow + pC.spacing

        if((crow+mrow)<(pg*pC.SizeY)):
            x.cell = 'A' + str(crow)
            crow = crow + mrow
        else:
            pg = pg + 1
            crow = (pg)*pC.SizeY + pC.tblock
            x.cell = 'A' + str(crow)
            crow = crow + mrow

    return crow,pg

def vCellAssign(varList,cRow,page):
    crow = cRow
    pg = page
    for x in varList:
        if (crow < (pg*pC.SizeY)):
            varList.cell = gCol(pC.cValue)+str(crow)
            crow = crow + 1
        else:
            pg = pg + 1
            crow = (pg)*pC.SizeY + pC.tblock
            varList.cell = gCol(pC.cValue)+str(crow)
            crow = crow + 1
    return crow, pg

def xlshlook(key,table,irow,lBool):
    tempformula = "=HLOOKUP("+key+","+table+","+irow+","+lBool+")"
    return tempformula

def xlsvlook(key,table,icol,lBool):
    tempformula = "=VLOOKUP("+key+","+table+","+icol+","+lBool+")"
    return tempformula

def ckfunction(condition,resultT,resultF):
    tempformula = "=IF("+condition+","+resultT,","+resultF,")"
    return tempformula

def processSym(String, varName, num):
    #convert to latex 
    eqn = simplify(String)
    lEqn = latex(eqn)

    #create the full equation
    lEqn = '$' + varName + '=' + lEqn + '$'

    #image filename & saving to file as *.png
    nPng = "seqEqn" + str(num)
    preview(lEqn, output = 'png', viewer = 'file', filename = nPng)

    return nPng, lEqn

def xlsProcessing():
    print("test")
    
while(True):
    print("Input your expression")
    key = raw_input(">>>")
    istring = xlsfuncFind(key)
    print(istring)
