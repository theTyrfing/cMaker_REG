#cMaker - Calculation Maker
#by Regan Lu

#Libaraies
import openpyxl as xl
from  Tkinter import *
import Tkinter, Tkconstants, tkFileDialog
from time import sleep

#Project Class
class project:
    title = ""
    pages = 0
    path = ""
    Ppath = ""
    Psheet = ""
    filename = ""
    prj = ""
    date = ""

#Calculation Class
class var: #variable class
    name = ""   #name
    var = ""    #variable
    value = ""  #value
    unit = ""   #unit
    note = ""   #notes
    ref = ""    #reference
    cell = ""   #location in sheet
    
    def intCont(self,ListV):
        self.name = ListV[0]
        self.var = ListV[1]
        self.value = ListV[2]
        self.unit = ListV[3]
        self.note = ListV[4]
        self.ref = ListV[5]
    
class seq: #sequence class
    seq = 0     #sequence order
    typ = 0     #type: 0 - formula, 1 - check/if statement, 2 - HLookup, 3 - Vlookup, 4 - Control variables, 5 - solver
    name = ""   #name
    var = ""    #variable
    unit = ""   #unit
    exp = ""    #expression
    lexp = ""   #latex expression
    lpng = ""   #latex path
    A = ""      #Control variable A
    B = ""      #Control variable B
    C = ""      #Control variable C
    ref = ""    #Reference
    cell = ""   #location in sheet
    xExp = ""   #Excel representation
    oCell = ""  #Over ride cell formula

    def intCont(self,listS):
        self.seq = listS[0]     #sequence order
        self.typ = listS[1]     #type: 0 - reg, 1 - iteration
        self.name = listS[2]    #name
        self.var = listS[3]     #variable
        self.unit = listS[4]    #unit
        self.exp = listS[5]     #expression
        self.A = listS[6]       #Control variable A
        self.B = listS[7]       #Control variable B
        self.C = listS[8]      #Control variable C
        self.ref = listS[9]    #Reference

class imp: #import database variable class
    name = ""   #Name of Database
    path = ""   #Path of Excel Database
    sheet = ""  #Sheet Name
    irange = "" #Range of table 
    cell = ""   #Location in sheet
    
    #Name of columns/rows for Python Only
    A = ""      #Control variable A
    Aunit = ""  #Control variable A unit
    B = ""      #Control variable B
    Bunit = ""  #Control variable B unit
    C = ""      #Control variable C
    Cunit = ""  #Control variable C unit
    D = ""      #Control variable D
    Dunit = ""  #Control variable D unit
    E = ""      #Control variable E
    Eunit = ""  #Control variable E unit
    F = ""      #Control variable F
    Funit = ""  #Control variable F unit
    
    def intCont(self,ListI):
        #Filled in info
        self.name = ListI[0]   #name
        self.var = ListI[1]    #variable
        self.path = ListI[2]   #path
        self.sheet = ListI[3]   #Sheet Name
        self.irange = ListI[4]  #Range of table         

        #Control Variables
        self.A = ListI[5]      #Control variable A
        self.B = ListI[6]      #Control variable B
        self.C = ListI[7]      #Control variable C
        self.D = ListI[8]      #Control variable D
        self.E = ListI[9]      #Control variable E
        self.F = ListI[10]      #Control variable F

        #"Missing" info
        self.Aunit = ListI[11]  #Control variable A unit
        self.Bunit = ListI[12]  #Control variable B unit
        self.Cunit = ListI[13]  #Control variable C unit
        self.Dunit = ListI[14]  #Control variable D unit
        self.Eunit = ListI[15]  #Control variable E unit
        self.Funit = ListI[16]  #Control variable F unit

#Variables
loaded = False
exiter = False
starter = False
importv = False
gFile = 0
Uproj = project()
    
#Functions
def menuOpt():
    print("================ MENU ================")
    print("Enter the following to options:")
    print("[I] to import list from drive")
    print("[S] to start generating new calculation sheet")
    print("[PV] to print variables list")
    print("[PS] to print sequence list")
    print("[PI] to print import list")
    print("[pJ] to print project details")
    print("[PJ] to edit project details")
    print("[E] to exit the program")
    print("[i] to show this menu again")
    print("================ MENU ================")

def addList(List, value):
    if isinstance(List,(list,)):
        temp = List
        temp = temp.append(value)
        return List
    else:
        return [value]

def intLister(num):
    if num == 0:
        return var()
    elif num == 1:
        return seq()
    elif num == 2:
        return imp()
    else:
        print("ERROR, Invaild Setting")
        return -1

def setList(num,vRange):
    tempList = 0
    mcol = vRange.max_column
    mrow = vRange.max_row - 1
    
    for row in vRange.iter_rows(min_row=1, row_offset=1,max_col=mcol,
                                max_row = mrow):
        TList = 0
        for cell in row:
            if(cell.value != None):
                TList = addList(TList,str(cell.value))
            else:
                TList = addList(TList,"")
        tmp = intLister(num)
        tmp.intCont(TList)

        tempList = addList(tempList,tmp)

    return tempList

def importList(varList,seqList,iList):
    #Return variables
    vList=0
    sList=0
    cList=0
    iiList=0
    root = Tk()
    
    #Get path of workbook
    gFile = tkFileDialog.askopenfilename(initialdir = "/",
                                                  title = "Select xlsx file to import",
                                                  filetypes = (("xlsx files","*.xlsx"),
                                                               ("xls files","*.xls"),
                                                               ("all files","*.*")))
    root.destroy()
    if gFile!=None:
        #Load Workbook & Ranges
        WB = xl.load_workbook(gFile)
        vRange = WB['Variables']
        sRange = WB['Sequences']
        iRange = WB['Import']

        #Load info into list
        vList=setList(0,vRange)
        sList=setList(1,sRange)
        iiList = setList(2,iRange)
        
    return vList,sList,iiList

#Startup message
def StartupMsg():
    print("======================================")
    print("WELCOME TO CMAKE")
    print("by Regan Lu")
    print("This program allow the user to create")
    print("calculation sheet with ease.")
    print("======================================")
    print("Note: it is recommended to define your") 
    print("variables outside of this program.")
    print("======================================")

    #Enter in project details at setup
    if(raw_input("Do you want to enter project details now? (Y/N): ") == "Y"):
        editProj(Uproj)
        print("Okay. Done!")
    menuOpt();

#Print current Lists functions
def printvarList(varList):
    print("Name | Variable | Value | Unit | Notes | Reference | Cell")
    print("=========================================================")
    for x in varList:
        print("%s | %s | %s | %s | %s | %s | %s" %(x.name, x.var, x.value,
                                                x.unit,x.note,x.ref,x.cell))
    
def printseqList(seqList):
    print("Sequence | Type | Name | Variable | Unit | Expression | LATEX | A | B | C | Notes | Reference | Cell")
    print("====================================================================================================")
    for x in seqList:
        print("%s | %s | %s | %s | %s | %s | %s | %s | %s | %s | %s | %s" %(x.seq, x.typ, x.name, x.var,
                                                                            x.exp, x.lexp, x.unit,x.A,x.B,
                                                                            x.C,x.ref,x.cell))
    
def printiList(iList):
    print("Name | Variable | Path | Sheet | Range | Cell | A | A_unit | B | B_unit | C | C_unit | D | D_unit | E | E_unit | F | F_unit")  
    print("===========================================================================================================================")
    for x in iList:
        print("%s | %s | %s | %s | %s | %s | %s | %s | %s | %s | %s | %s | %s | %s | %s | %s | %s | %s" %(x.name, x.var, x.path, x.sheet, x.irange,
                                                                                                          x.cell, x.A, x.Aunit,x.B,x.Bunit,
                                                                                                          x.C,x.Cunit,x.D,x.Dunit,
                                                                                                          x.E, x.Eunit,x.F,x.Funit))
def printProj(proj):
    print("Project Name: %s" % proj.prj)
    print("Project Pages: %s" % proj.pages)
    print("Project Date: %s" % proj.date)
    print("Project Path: %s" % proj.path)
    print("Project Paper Path: %s" % proj.Ppath)
    print("Project Sheet Name: %s" % proj.Psheet)
    print("Project filename: %s" % proj.filename)

def editProj(proj):
    root = Tk()
    proj.prj = raw_input("Project Name:")
    proj.date = raw_input("Project Date:")
    proj.Psheet = raw_input("Project Sheet Name:")
    proj.filename = raw_input("Project filename:")
    
    print('Please select path for workbook.')
    sleep(1)
    proj.path = tkFileDialog.askdirectory(initialdir = "/", title = "Workbook Path") +"/"
    print('Please select path for template.')
    sleep(1)
    proj.Ppath = tkFileDialog.askopenfilename(initialdir = "/",
                                                  title = "Select xlsx file to paper path",
                                                  filetypes = (("xlsx files","*.xlsx"),
                                                               ("xls files","*.xls"),
                                                               ("all files","*.*")))
    root.destroy()

#Swtich Cases for User Input
def uiInput(Key,varList,seqList,iList,proj):
    if(Key == 'I'):    #I for Import Lists
        return False,False,True
    elif(Key == 'E'):  #E for Exit
        return False,True,False
    elif(Key == 'S'):  #S for Start
        return True,False,False
    elif(Key == 'PV'): #PV for print Variable List
        printvarList(varList)
        return False,False,False
    elif(Key == 'PS'): #PS for print Sequence List 
        printseqList(seqList)
        return False,False,False
    elif(Key == 'PI'): #PS for print Sequence List 
        printiList(iList)
        return False,False,False
    elif(Key == 'pJ'): #Print Project Details
        printProj(proj)
        return False,False,False
    elif(Key == 'PJ'): #Edit Project Details
        editProj(proj)
        return False,False,False
    elif(Key == 'i'):  #i for print this menu
        menuOpt()
        return False,False,False
    else:
        print("Error: Invalid Input")
        return False,False,False

#Main Loop
#Start up
StartupMsg()

#Storage for imports
varList = 0
seqList = 0
ckList = 0
iList = 0

while(exiter == False):
    key = raw_input(">>>")
    starter,exiter,importv = uiInput(key,varList,seqList,iList,Uproj)

    if(importv):
        varList,seqList,iList = importList(varList,seqList,iList)

    if(starter):
        Uproj.pages = xlsPP(varList,seqList,iList,Uproj.prj)
        WorkBook, Sheet = createNew(Uproj)
        xlsProcessing(varList, seqList, iList, Sheet)
        WorkBook.save(filename = pj.path + pj.filename)
