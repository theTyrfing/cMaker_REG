#Equation Parser for python and Excel code
#Function Libaray
#by Regan Lu

#import libaray functions
from math import *
from sympy import simplify, sstrrepr, latex, preview, Symbol 
import pageConfig as pC
from openpyxl.utils.cell import _get_column_letter as gCol
from openpyxl.utils.cell import coordinate_to_tuple as ctt
from openpyxl import load_workbook
from os.path import exists, dirname, realpath
from os import makedirs

#local variables
logicOpt = ["==",">","<",">=","<=","!="]
xlsLogic = ["=",">","<",">=","<=","!="]

#defining the excel function list
pyNameList = ['fabs(','factorial(','ceil(','log(','pow(','**' ]
xlsNameList = ['abs(','fact(','ceiling(','ln(','power(','^']

#Functions

def bForm(String):
    
    tString = '=' + String

    return tString

#convert equation from processed sympy equation
#to excel friendly equation
def xlsfuncFind(String):
    #temp string
    tString = String
    
    #index for names
    i = 0
    
    #Excel Function Replacement Loop
    #Limited libaray replacement, further development will be required
    #Should be good enough for most application anyways
    for x in pyNameList:
        tString = tString.replace(x,xlsNameList[i])
        i = i + 1
    
    return tString

#Replace variable with excel reference
def xlsCell(varList, seqList, String):
    #temp string
    tString = String

    #Variable Replacement loops
    #variable list
    for x in varList:
        tString = tString.replace(x.var,x.cell)

    #sequence list
    for x in seqList:
        tString = tString.replace(x.var,x.cell)
    
    return tString

def overRider(cCell):
    oTuple = ctt(cCell)
    row = oTuple[1]

    oCell = pC.colForm + str(row + pC.rOver)
    fCell = pC.colForm + str(row + pC.rForm)
    oRider = ckfunction(oCell+'=\"\"',fCell,oCell)

    return oRider

def Assigner(varList, seqList, iList):
    cRow = pC.tblock + pC.SizeY + 1
    page = 1

    cRow,page = vCellAssign(varList,cRow,page)
    cRow,page = iCellAssign(iList,cRow,page)
    cRow,page = sCellAssign(seqList,cRow,page)

    return page

def sCellAssign(seqList,cRow,page):
    crow = cRow
    pg = page

    for x in seqList:
        crow = crow + pC.spacing

        if(x.type == '0'):
            if((crow+fSize)<(pg*pC.SizeY)):
               x.cell = pC.colForm + str(crow-pC.rEqt)
               crow = crow+fSize
            else:
                pg = pg + 1
                crow = (pg)*pC.SizeY + pC.tblock
                x.cell = pC.colForm + str(crow-pC.rEqt)
                crow = crow + pC.fSize
        elif(x.type == '1'):
            if((crow+ifSize)<(pg*pC.SizeY)):
               x.cell = pC.colForm + str(crow-pC.rift)
               crow = crow+fSize
            else:
                pg = pg + 1
                crow = (pg)*pC.SizeY + pC.tblock
                x.cell = pC.colForm + str(crow-pC.rift)
                crow = crow + pC.ifSize
        elif(x.type == '2'):
            if((crow+ifSize)<(pg*pC.SizeY)):
               x.cell = pC.colForm + str(crow-pC.rift)
               crow = crow+fSize
            else:
                pg = pg + 1
                crow = (pg)*pC.SizeY + pC.tblock
                x.cell = pC.colForm + str(crow-pC.rift)
                crow = crow + pC.ifSize
        elif(x.type == '3'):
            if((crow+ifSize)<(pg*pC.SizeY)):
               x.cell = pC.colForm + str(crow-pC.rift)
               crow = crow+fSize
            else:
                pg = pg + 1
                crow = (pg)*pC.SizeY + pC.tblock
                x.cell = pC.colForm + str(crow-pC.rift)
                crow = crow + pC.ifSize
        else:
            print("Error: Incorret Type")
            
def iCellAssign(iList,cRow,page):
    crow = cRow
    pg = page
    
    for x in iList:
        path = x.path
        f = load_workbook(path)
        mrow = max_row(f.active)
        crow = crow + pC.spacing

        if((crow+mrow)<(pg*pC.SizeY)):
            x.cell = pC.colTable + str(crow)
            crow = crow + mrow
        else:
            pg = pg + 1
            crow = (pg)*pC.SizeY + pC.tblock
            x.cell = pC.colTable + str(crow)
            crow = crow + mrow

    return crow,pg

def vCellAssign(varList,cRow,page):
    crow = cRow
    pg = page
    for x in varList:
        if (crow < (pg*pC.SizeY)):
            varList.cell = gCol(pC.cValue)+str(crow)
            crow = crow + 1
        else:
            pg = pg + 1
            crow = (pg)*pC.SizeY + pC.tblock
            varList.cell = gCol(pC.cValue)+str(crow)
            crow = crow + 1
    return crow, pg

def findTable(iList, varg):
    index = 0 
    for im in iList:
        if(im.var == varg):
            return index
        else:
            index = index + 1

    return -1
    
def xlshlook(key,table,irow,lBool):
    tempformula = "=HLOOKUP("+key+","+table+","+irow+","+lBool+")"
    return tempformula

def xlsvlook(key,table,icol,lBool):
    tempformula = "=VLOOKUP("+key+","+table+","+icol+","+lBool+")"
    return tempformula

def ckfunction(condition,resultT,resultF):
    tcond = condition
    tcond = tcond.replace("==","=")
    tcond = temp.replace("!=","<>")

    tempformula = "=IF("+tcond+","+resultT,","+resultF,")"
    return tempformula

def processSym(String, varName, num, pj):
    #convert to latex 
    eqn = simplify(String)
    lEqn = latex(eqn)

    #create the full equation
    lEqn = '$' + varName + '=' + lEqn + '$'

    #Creating directory if it does not have one
    pngPath = dirname(realpath(__file__)) + "/" + pj + "/"
    if not exists(pngPath):
        os.makedirs(pngPath)

    #image filename & saving to file as *.png    
    nPng = pngPath + "seqEqn" + str(num)
    preview(lEqn, output = 'png', viewer = 'file', filename = nPng)

    return nPng, lEqn

def ckLatex(condition, num, pj):
    for x in logicOpt:
        index = condition.find(x)
        temp = condition
        off = len(x) - 1
        
        if(index != -1):
            #Split into two parts
            pt1 = simplify(condition[0:index])
            pt2 = simplify(condition[(index+off):len(condition)])

            #Latex both and combine
            pt1 = latex(pt1)
            pt2 = latex(pt2)
            cktext = '$' + pt1 + x + pt2 + '$'

            #Creating directory if it does not have one
            pngPath = dirname(realpath(__file__)) + "/" + pj + "/"
            if not exists(pngPath):
                os.makedirs(pngPath)

            #image filename & saving to file as *.png    
            nPng = pngPath + "seqEqn" + str(num)
            preview(cktext, output = 'png', viewer = 'file', filename = nPng)

            return cktext, nPng
        
        else:
            print("Error: Invalid Expression")
            return -1, -1
    
def xlsPP(varList,seqList,iList,prj):
    #Getting name of project
    if(prj == ""):
        dprj = "eqn-dir"
    else:
        dprj = prj

    #Assigning cell reference to all variables, sequences, and tables
    pages = Assigner(varList,seqList,iList)

    #Sequence formula loop
    for sq in seqList:
        if(sq.typ == 0):
            
            #Latex and overrider formulas
            sq.lpng, sq.lexp = processSym(sq.exp,sq.var,sq.seq,dprj)
            sq.oCell = overRider(sq.cell)

            #Configuring formula for xls
            tmpFormula = sq.exp
            tmpFormula = xlsCell(varList, seqList,tmpFormula)
            tmpFormula = xlsfuncFind(tmpFormula)
            tmpFormula = bForm(tmpFormula)

            sq.xExp = tmpFormula
            
        elif(sq.typ == 1):
            sq.lpng, sq.lexp = ckLatex(sq.exp)
            sq.xExp = ckfunction(sq.exp,sq.A,sq.B)
            
        elif(sq.typ == 2):
            tab = findTable(iList,sq.A)
            sq.xExp = xlshlook(sq.exp,tab,sq.A,sq.B)

        elif(sq.typ == 3):
            tab = findTable(iList,sq.A)
            sq.xExp = xlsvlook(sq.exp,tab,sq.A,sq.B)

        else:
            print("Error: invalid type")
    

    return pages
