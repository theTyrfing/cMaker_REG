#Excel Compiler module Module for cMaker
#by Regan Lu

#Libaraies
from openpyxl import load_workbook, Workbook
from copy import copy
from pageConfig import SizeX, SizeY

#page class
class Pages:
    row1 = 0 
    col1 = 0
    row2 = 0
    col2 = 0

#Functions
def pageAssign(page):
    tmp = 0
    beginRow = 1
    endRow = 1
    
    for x in range(0,page):
        if tmp == 0:
            tmp = Pages()
            tmp.row1 = 1
            tmp.row2 = SizeY
            tmp.col1 = 1
            tmp.col2 = SizeX
            tmp = [tmp]
            endRow = SizeY
        else:
            beginRow = endRow + beginRow
            endRow = endRow + SizeY
            iTmp = Pages()
            iTmp.row1 = beginRow
            iTmp.row2 = endRow
            iTmp.col1 = 1
            iTmp.col2 = SizeX
            tmp = tmp.append(iTmp)

    return tmp

def mergedCells(paper):
    mRanges = paper.merged_cells.ranges()
    tmp = 0

    for x in mRanges:
        if tmp == 0:
            tmp = x.coord
            tmp = [tmp]
        else:
            iTmp = x.coord
            tmp = tmp.append(iTmp)

    return tmp

def pageMaker(path,shtName,lPages,wSheet):
    #Template Copy open to work on
    wkTmp = load_workbook(path)
    paper = wkTmp[shtName]
    mRange = mergedCells(paper)

    wSheet.target.sheet_format = copy(paper.sheet_format)
    wSheet.sheet_properties = copy(paper.sheet_properties)

    for m in mRange:
         wSheet.merge_cells(m)
    
    for page in lPages:
        #iterate rows to modify
        for i in range(page.col1,page.col2):
            for j in range(page.row1,page.row2):
                 #Getting Cells
                 tCell = wSheet.cell(col = i, row = j)
                 diff = j - page.row1 - 1
                 sCell = paper.cell(col = i, row = diff)

                 tCell.value = sCell.value
                 tCell.datatype = sCell.datatype

                 if sCell.has_style:
                     tCell._style = copy(sCell._style)

                 if sCell.hyperlink:
                     tCell._hyperlink = copy(sCell.hyperlink)

                 if source_cell.comment:
                     tCell.comment = copy(sCell.comment)


def createNew(pj):
    #Create new workbook
    aPage = pageAssign(pj.page)
    eWB = Workbook()
    eSheet = eWB.title(pj.title)

    #Saving workbook
    tmpPath = pj.path + pj.filename
    eWB.save(filename = tmpPath)

    #Setup Sheet
    pager = pageAssign(pj.pages)
    pageMaker(pj.Ppath,pj.Psheet,pager,eSheet)

    #return new workbook
    return eWB,eSheet

def xlsProcessing(varList, seqList, iList, eSheet):
    print("do nothing")

